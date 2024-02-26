import xml.etree.ElementTree as ET
import openpyxl

# Parse Robot Framework XML output file
tree = ET.parse(r'C:\Users\MSI\Desktop\Studies\PFE\Automation-Test-Using-Robot-Framework\VS_codes\tests_Robot_Framework\OutputChangeHostname_WebInterface.xml')
root = tree.getroot()

# Extract information from XML file
test_results = []
for test in root.findall('.//test'):
    test_name = test.attrib['name']
    for kw in test.findall('.//kw'):
        kw_name = kw.attrib['name']
        status = kw.find('.//status').attrib['status']
        error = kw.find('.//msg').text if status == 'FAIL' else ''
        test_results.append((test_name, kw_name, status, error))

# Create or load Excel workbook
wb = openpyxl.Workbook()
ws = wb.active

# Populate Excel worksheet with test results
for row, result in enumerate(test_results, start=1):
    ws.cell(row=row, column=1, value=result[0])
    ws.cell(row=row, column=2, value=result[1])
    ws.cell(row=row, column=3, value=result[2])
    ws.cell(row=row, column=4, value=result[3])
# Save Excel workbook
wb.save(r'C:\Users\MSI\Desktop\Studies\PFE\Automation-Test-Using-Robot-Framework\VS_codes\tests_Robot_Framework\ChangeHostname_WebInterface_results.xlsx')
